import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_dispatch_excel(input_filename: str, output_filename: str = '쿠팡로켓_출고리스트_완성.xlsx'):
    try:
        df = pd.read_excel(input_filename)
        print(f"'{input_filename}' 파일을 성공적으로 읽어왔습니다.")
    except FileNotFoundError:
        print(f"경고: '{input_filename}' 파일을 찾을 수 없어 빈 데이터프레임으로 시작합니다.")
        df = pd.DataFrame(columns=['출고 모델명', '단품명', '출고요청', '배송지', '출고예정일'])

    # 컬럼이 없는 경우 빈 문자열로 초기화
    for col in ['출고 모델명', '단품명', '출고요청', '배송지', '출고예정일']:
        if col not in df.columns:
            df[col] = ''

    # 1. 정렬: 배송지(오름차순) -> 출고 모델명(오름차순) -> 단품명(커스텀)
    size_order = ['3XS', '2XS', 'XS', 'S', 'M', 'L', 'XL', '2XL', '3XL', '4XL']
    
    df['단품명'] = df['단품명'].astype(str)
    df['단품명_순위'] = df['단품명'].apply(lambda x: size_order.index(x.upper()) if x.upper() in size_order else 999)

    # 숫자형 변환 및 정렬을 위한 빈값 처리
    df['출고요청'] = pd.to_numeric(df['출고요청'], errors='coerce').fillna(0).astype(int)
    df['배송지'] = df['배송지'].fillna('')
    df['출고 모델명'] = df['출고 모델명'].fillna('')
    
    df = df.sort_values(by=['배송지', '출고 모델명', '단품명_순위'], ascending=[True, True, True])

    # 총합 계산
    total_request = df['출고요청'].sum()

    # 날짜 추출
    date_str = '0000-00-00'
    if not df['출고예정일'].empty:
        first_valid = df['출고예정일'].replace('', pd.NA).dropna().iloc[0] if not df['출고예정일'].replace('', pd.NA).dropna().empty else None
        if first_valid:
            try:
                date_str = pd.to_datetime(first_valid).strftime('%Y-%m-%d')
            except:
                date_str = str(first_valid)[:10]

    # Openpyxl 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "로켓출고리스트"

    # 스타일 설정
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    fill_pink = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    fill_light_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    font_dark_red = Font(color="8B0000", bold=True)
    
    # 1행: 쿠팡 로켓 출고 리스트
    cell_b1 = ws.cell(row=1, column=2, value="쿠팡 로켓 출고 리스트")
    cell_b1.font = Font(size=28, bold=True)
    cell_b1.alignment = align_center
    ws.merge_cells("B1:E1")

    # 2행: 출고일 및 총합
    cell_b2 = ws.cell(row=2, column=2, value=f"출고일: {date_str}")
    cell_b2.font = Font(size=11, bold=True)
    cell_b2.alignment = align_right
    ws.merge_cells("B2:C2")

    cell_d2 = ws.cell(row=2, column=4, value=total_request)
    cell_d2.font = Font(size=11, bold=True)
    cell_d2.alignment = align_center

    # 3행: 헤더 (핑크 배경)
    headers = ['출고 모델명', '단품명', '출고요청', '배송지']
    for idx, header in enumerate(headers, start=2): # B(2), C(3), D(4), E(5)
        cell = ws.cell(row=3, column=idx, value=header)
        cell.font = Font(size=11, bold=True)
        cell.alignment = align_center
        cell.border = border_thin
        cell.fill = fill_pink

    # 4행부터 데이터 추가 및 G열 중복 검사 준비
    seen_combinations = set()
    duplicate_combinations = set()

    for idx, (_, row) in enumerate(df.iterrows(), start=4):
        model_out = row['출고 모델명']
        size_name = row['단품명']
        req_qty = row['출고요청']
        location = row['배송지']

        ws.cell(row=idx, column=2, value=model_out).alignment = align_center
        ws.cell(row=idx, column=3, value=size_name).alignment = align_center
        ws.cell(row=idx, column=4, value=req_qty).alignment = align_center
        ws.cell(row=idx, column=5, value=location).alignment = align_center

        for c in range(2, 6):
            ws.cell(row=idx, column=c).border = border_thin

        # G열: 출고 모델명_단품명
        combo_str = f"{model_out}_{size_name}"
        ws.cell(row=idx, column=7, value=combo_str).alignment = align_center
        ws.cell(row=idx, column=7).border = border_thin
        
        if combo_str in seen_combinations:
            duplicate_combinations.add(combo_str)
        else:
            seen_combinations.add(combo_str)

    # 중복행 조건부 서식 적용 (연한 빨강 채우기 및 진한 텍스트)
    for r in range(4, 4 + len(df)):
        combo_val = ws.cell(row=r, column=7).value
        if combo_val in duplicate_combinations:
            ws.cell(row=r, column=7).font = font_dark_red
            ws.cell(row=r, column=7).fill = fill_light_red

    # 배송지 컬럼(E열, col=5) 자동 병합 처리
    if len(df) > 0:
        start_row = 4
        current_val = ws.cell(row=start_row, column=5).value
        
        for r in range(5, 4 + len(df)):
            val = ws.cell(row=r, column=5).value
            if val != current_val:
                # 값이 달라지면 이전 덩어리 병합 (2줄 이상일 때만)
                if r - 1 > start_row:
                    ws.merge_cells(start_row=start_row, start_column=5, end_row=r-1, end_column=5)
                start_row = r
                current_val = val
        
        # 마지막 그룹 병합 처리
        if (3 + len(df)) > start_row:
            ws.merge_cells(start_row=start_row, start_column=5, end_row=3 + len(df), end_column=5)

    # 컬럼 너비 조정
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['G'].width = 30

    wb.save(output_filename)
    print(f"=== 엑셀 파일 생성 완료: {output_filename} ===")

if __name__ == "__main__":
    create_dispatch_excel(input_filename='출고요청.xlsx')
