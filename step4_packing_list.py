import streamlit as st
import pandas as pd
import io
import json
import traceback
import google.generativeai as genai
from PIL import Image

def init_gemini():
    try:
        api_key = st.secrets["GEMINI_API_KEY"]
        genai.configure(api_key=api_key)
        
        # 무료 요금제에서 하루 20건 제한이 있는 3.x 최신 모델 대신,
        # 하루 1500건까지 무료로 넉넉하게 쓸 수 있는 2.0 버전 모델을 최우선으로 배치합니다.
        model_names = [
            'gemini-2.0-flash',
            'gemini-2.0-flash-lite',
            'gemini-2.5-flash',
            'gemini-flash-lite-latest'
        ]
        return model_names
    except Exception as e:
        return None

def extract_info_from_images_batch(model_names, image_bytes_list):
    try:
        imgs = [Image.open(io.BytesIO(b)) for b in image_bytes_list]
        num_imgs = len(imgs)
        
        prompt = f"""
        입력된 사진들({num_imgs}장)에는 택배 박스 겉면을 촬영한 사진들이 순서대로 있습니다.
        각 사진(박스) 안에는 **여러 장의 작은 상품 라벨 스티커들**과 **1개의 큰 박스 번호(숫자)**가 있습니다.
        
        당신의 임무는 각 사진에 있는 '모든' 상품 라벨 스티커 정보를 하나도 빠짐없이 찾아내어 JSON 배열로 만드는 것입니다.
        (주의: 한 장의 사진에 상품 라벨이 9개 붙어있다면, 반드시 9개의 JSON 객체를 만들어야 합니다. 절대 누락하지 마세요!)

        추출할 정보:
        1. destination (배송지): 괄호 () 안에 적힌 텍스트 (예: 시흥2)
        2. model_name (모델명): 영어와 숫자가 혼합된 상품코드 (예: TM-FSS44-ZZNVY)
        3. item_name (단품명/사이즈): 수량 대괄호 앞의 텍스트 (예: XL, 2XL, M 등)
        4. qty (수량): 대괄호 [] 안에 적힌 숫자 (예: 16)
        5. box_no (박스번호): 작은 상품 라벨들과는 별도로, 박스에 '아주 크게 인쇄된 종이'가 붙어있거나 매직으로 크게 적힌 1개의 대표 숫자 (예: 5). 
           ★중요★ 해당 사진에서 찾은 모든 상품 라벨 결과물(`box_no`)에는 그 사진의 대표 박스 번호가 동일하게 들어가야 합니다!

        결과는 순수 JSON '배열(Array)' 형식으로만 반환하세요. 텍스트나 다른 설명은 절대 추가하지 마세요.
        [
          {{
            "destination": "시흥2",
            "model_name": "TM-FSS44-ZZNVY",
            "item_name": "XL",
            "qty": 16,
            "box_no": "5"
          }},
          {{
            "destination": "시흥2",
            "model_name": "TM-FSS44-ZZWHT",
            "item_name": "M",
            "qty": 14,
            "box_no": "5"
          }}
        ]
        """
        
        last_error = None
        for m_name in model_names:
            try:
                model = genai.GenerativeModel(m_name)
                # 프롬프트 텍스트 다음에 여러 개의 이미지를 함께 리스트로 묶어서 전달
                content_parts = [prompt] + imgs
                response = model.generate_content(content_parts)
                text = response.text.strip()
                
                if text.startswith("```json"):
                    text = text[7:]
                if text.startswith("```"):
                    text = text[3:]
                if text.endswith("```"):
                    text = text[:-3]
                    
                data = json.loads(text.strip())
                return data
            except Exception as e:
                last_error = e
                continue
                
        # 모든 모델 실패 시 마지막 에러 발생 (재시도 로직이 잡을 수 있도록)
        raise Exception(f"{last_error}")
        
    except Exception as e:
        # 상위 루프에서 429 에러 등을 캐치할 수 있도록 다시 raise
        raise e

def match_and_generate_packing_list(df_req, ocr_results):
    df_req.columns = df_req.columns.astype(str).str.strip().str.replace('\n', '')
    
    col_mapping = {
        '발주번호': ['발주번호', '발주 번호', '발주서번호'],
        '배송지': ['배송지', '도착지', '배송처'],
        '단품명': ['단품명', '상품명', '품명'],
        '상품바코드': ['상품바코드', '바코드', '바코드번호'],
        '출고모델명': ['출고모델명', '출고 모델명', '모델명'],
        '출고요청': ['출고요청', '출고요청수량', '수량', '요청수량']
    }
    
    actual_cols = {}
    for req_key, possible_names in col_mapping.items():
        found = False
        for p_name in possible_names:
            if p_name in df_req.columns:
                actual_cols[req_key] = p_name
                found = True
                break
        if not found:
            st.error(f"출고요청 엑셀 파일에 '{req_key}' 역할을 하는 컬럼이 없습니다.")
            st.warning(f"현재 엑셀에서 인식된 컬럼들: {list(df_req.columns)}")
            return None
            
    df = df_req.copy()
    df.rename(columns={actual_cols[k]: k for k in col_mapping.keys()}, inplace=True)
    
    df['박스번호'] = ''
    df['출고요청'] = pd.to_numeric(df['출고요청'], errors='coerce').fillna(0).astype(int)
    
    matched_indices = set()
    
    for item in ocr_results:
        box_no = str(item.get('box_no', '')).strip()
        dest = str(item.get('destination', '')).strip().lower().replace(' ', '')
        model_name = str(item.get('model_name', '')).strip().lower().replace(' ', '')
        item_name = str(item.get('item_name', '')).strip().lower().replace(' ', '')
        qty = item.get('qty', 0)
        
        try:
            qty = int(qty)
        except:
            continue
            
        if not box_no:
            continue
            
        for idx, row in df.iterrows():
            if idx in matched_indices:
                continue
                
            excel_dest = str(row['배송지']).strip().lower().replace(' ', '')
            excel_model = str(row['출고모델명']).strip().lower().replace(' ', '')
            excel_item = str(row['단품명']).strip().lower().replace(' ', '')
            excel_qty = row['출고요청']
            
            if excel_qty == qty and excel_model == model_name and excel_item == item_name and (dest in excel_dest or excel_dest in dest):
                df.at[idx, '박스번호'] = box_no
                matched_indices.add(idx)
                break
                
    out_df = pd.DataFrame()
    out_df['발주번호'] = df['발주번호']
    out_df['출고모델명'] = df['출고모델명']
    out_df['단품명'] = df['단품명']
    out_df['수량'] = df['출고요청']
    out_df['박스번호'] = df['박스번호']
    
    out_df['팔렛트번호'] = ''
    out_df['배송지'] = df['배송지']
    out_df['상품바코드'] = df['상품바코드']
    
    # 중복여부 파악 (배송지, 출고모델명, 단품명이 동일한 건)
    out_df['중복여부'] = out_df.duplicated(subset=['배송지', '출고모델명', '단품명'], keep=False)
    
    temp_box_num = pd.to_numeric(df['박스번호'], errors='coerce').fillna(999999).astype(int)
    out_df['sort_box'] = temp_box_num.astype(str).str.zfill(10)
    
    out_df = out_df.sort_values(by=['sort_box', '배송지']).drop(columns=['sort_box'])
    out_df = out_df.reset_index(drop=True)
    
    return out_df

def generate_excel_bytes(df, ocr_results):
    output = io.BytesIO()
    from openpyxl.styles import PatternFill
    dup_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: 패킹리스트
        df_out = df.drop(columns=['중복여부'])
        df_out.to_excel(writer, index=False, sheet_name='패킹리스트')
        ws1 = writer.sheets['패킹리스트']
        
        ws1.cell(row=1, column=9, value="매핑키")
        for idx, row in df.iterrows():
            row_num = idx + 2
            ws1.cell(row=row_num, column=9, value=f'=A{row_num}&"_"&H{row_num}')
            
            if row['중복여부']:
                for col in range(1, 10):
                    ws1.cell(row=row_num, column=col).fill = dup_fill
                    
        for col_letter, width in zip(['A','B','C','D','E','F','G','H','I'], [15, 20, 15, 8, 10, 15, 20, 15, 25]):
            ws1.column_dimensions[col_letter].width = width
            
        # Sheet 2: AI_판독결과
        df_ocr = pd.DataFrame(ocr_results)
        if not df_ocr.empty:
            df_ocr = df_ocr.rename(columns={'destination': '배송지', 'model_name': '모델명', 'item_name': '단품명', 'qty': '수량', 'box_no': '박스번호'})
            cols = [c for c in ['배송지', '모델명', '단품명', '수량', '박스번호'] if c in df_ocr.columns]
            df_ocr = df_ocr[cols]
        else:
            df_ocr = pd.DataFrame(columns=['배송지', '모델명', '단품명', '수량', '박스번호'])
            
        df_ocr.to_excel(writer, index=False, sheet_name='AI_판독결과')
        ws2 = writer.sheets['AI_판독결과']
        for col_letter, width in zip(['A','B','C','D','E'], [20, 20, 15, 10, 15]):
            ws2.column_dimensions[col_letter].width = width
            
    return output.getvalue()

def render_packing_list_page():
    st.subheader("📦 패킹리스트 파일 생성 (AI 자동인식)")
    st.markdown("출고요청파일 엑셀과 박스 패킹 사진을 업로드하면, AI 비전이 사진의 **박스번호**와 **라벨(모델명/수량)**을 자동으로 판독하여 엑셀과 매칭해 줍니다!")
    
    gemini_model = init_gemini()
    if not gemini_model:
        st.error("⚠️ 구글 Gemini API 키가 설정되지 않았거나 올바르지 않습니다. 스트림릿 Secrets에 `GEMINI_API_KEY`를 설정해주세요.")
        return
        
    with st.expander("🛠️ API 모델 연결 테스트 (디버깅용)"):
        if st.button("사용 가능한 AI 모델 목록 불러오기"):
            try:
                available_models = []
                for m in genai.list_models():
                    if 'generateContent' in m.supported_generation_methods:
                        available_models.append(m.name)
                if available_models:
                    st.success("API 연결 성공! 사용 가능한 모델들:")
                    st.write(available_models)
                else:
                    st.warning("사용 가능한 모델이 하나도 없습니다. API 키 설정이나 구글 계정 권한을 확인해주세요.")
            except Exception as e:
                st.error(f"모델 목록 불러오기 실패: {e}")

    req_file = st.file_uploader("1. 출고요청파일 업로드 (Excel)", type=['xlsx'])
    img_files = st.file_uploader("2. 박스 패킹 사진 다중 업로드 (JPG/PNG)", type=['jpg', 'jpeg', 'png'], accept_multiple_files=True)
    
    if req_file and img_files:
        if st.button("🚀 패킹리스트 AI 생성 시작"):
            with st.spinner("엑셀 데이터를 분석하고 AI가 사진들을 묶음(Batch) 판독 중입니다... (속도가 매우 빠릅니다!)"):
                try:
                    df_req = pd.read_excel(req_file)
                    
                    all_ocr_results = []
                    progress_bar = st.progress(0)
                    total_imgs = len(img_files)
                    
                    import time
                    
                    # 5장씩 묶어서 처리
                    BATCH_SIZE = 5
                    for i in range(0, total_imgs, BATCH_SIZE):
                        batch_files = img_files[i:i+BATCH_SIZE]
                        img_bytes_list = [f.read() for f in batch_files]
                        
                        max_retries = 3
                        retry_count = 0
                        res = None
                        
                        while retry_count < max_retries:
                            try:
                                res = extract_info_from_images_batch(gemini_model, img_bytes_list)
                                break
                            except Exception as e:
                                err_str = str(e)
                                if "429" in err_str or "exceeded" in err_str.lower() or "quota" in err_str.lower():
                                    retry_count += 1
                                    st.warning(f"API 요청 과부하 감지. 15초 대기 후 묶음 재시도합니다... ({retry_count}/{max_retries})")
                                    time.sleep(15.0)
                                else:
                                    st.error(f"이미지 판독 중 오류 발생: {err_str}")
                                    break
                                    
                        if res:
                            all_ocr_results.extend(res)
                            
                        # 프로그레스 바 업데이트
                        current_processed = min(i + BATCH_SIZE, total_imgs)
                        progress_bar.progress(current_processed / total_imgs)
                        
                        # 배치 처리 후 무료 API 분당 제한(15회) 방지를 위해 아주 짧게 1.5초만 대기
                        # 30장을 처리해도 배치 6번이면 끝나므로 분당 제한에 걸릴 위험이 사실상 없음
                        if current_processed < total_imgs:
                            time.sleep(1.5)
                            
                    st.success(f"총 {total_imgs}장의 사진에서 {len(all_ocr_results)}개의 라벨 정보를 판독했습니다!")
                    
                    final_df = match_and_generate_packing_list(df_req, all_ocr_results)
                    if final_df is not None:
                        excel_data = generate_excel_bytes(final_df, all_ocr_results)
                        
                        st.success("패킹리스트 생성이 완료되었습니다! 아래 버튼을 눌러 다운로드하세요.")
                        st.download_button(
                            label="📥 패킹리스트_완성.xlsx 다운로드",
                            data=excel_data,
                            file_name="패킹리스트_완성.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        st.markdown("### 📊 최종 결과 미리보기")
                        st.dataframe(final_df.drop(columns=['중복여부']))
                        
                except Exception as e:
                    st.error(f"처리 중 오류가 발생했습니다: {e}")
                    st.error(traceback.format_exc())
